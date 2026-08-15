#!/usr/bin/env python3
import json
import pathlib
import sys
import zipfile
from openpyxl import load_workbook

from archive_package import archive_index, verify_archives

root = pathlib.Path(sys.argv[1])
files = [p for p in root.rglob("*") if p.is_file()]
change_lists = [p for p in files if ("变更清单" in p.name or "核对清单" in p.name) and p.suffix.lower() == ".xlsx"]
revised = [p for p in files if p.suffix.lower() in {".docx", ".xlsx", ".zip"} and p not in change_lists]
errors = []
if not change_lists:
    errors.append("missing archive change/checklist XLSX")
if not revised:
    errors.append("missing revised archive output")
task_path = root.parent / "task.json"
archive_kind = "personnel"
task = {}
if task_path.exists():
    try:
        task = json.loads(task_path.read_text(encoding="utf-8"))
        archive_kind = ((task.get("payload") or {}).get("archiveKind") or "personnel")
    except (OSError, json.JSONDecodeError):
        errors.append("unreadable task.json")
expected_headers = (
    ["环节", "必备项目", "依据来源", "状态", "冲突", "备注"]
    if archive_kind == "trainingPeriod"
    else ["人员", "来源文件", "字段", "原值", "新值", "依据", "状态", "备注"]
)
for file_path in change_lists:
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=False)
        worksheet = workbook.active
        actual_headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        if actual_headers[:len(expected_headers)] != expected_headers:
            errors.append(f"incorrect Chinese checklist headers: {file_path.name}")
    except (OSError, ValueError, StopIteration, zipfile.BadZipFile):
        errors.append(f"unreadable checklist workbook: {file_path.name}")
for file_path in revised + change_lists:
    if file_path.suffix.lower() in {".docx", ".xlsx", ".zip"}:
        try:
            with zipfile.ZipFile(file_path) as archive:
                if not archive.namelist():
                    errors.append(f"empty package: {file_path.name}")
        except (OSError, zipfile.BadZipFile):
            errors.append(f"unreadable package: {file_path.name}")
for file_path in [item for item in revised if item.suffix.lower() == ".zip"]:
    try:
        archive_index(file_path, include_hashes=False)
    except (OSError, ValueError, zipfile.BadZipFile) as error:
        errors.append(f"unsafe or invalid archive ZIP: {file_path.name}: {error}")

archive_upload = task.get("archive_upload") or {}
if archive_upload:
    source_zips = [
        pathlib.Path(str(item.get("stored_path") or ""))
        for item in (task.get("source_files") or [])
        if pathlib.Path(str(item.get("original_name") or "")).suffix.lower() == ".zip"
    ]
    output_zips = [item for item in revised if item.suffix.lower() == ".zip"]
    if len(source_zips) != 1:
        errors.append("archive task must contain one source ZIP")
    if not output_zips:
        errors.append("archive task must produce a revised ZIP")
    if len(source_zips) == 1 and output_zips:
        try:
            report = verify_archives(source_zips[0], output_zips[0])
            (root / "档案保真核验.json").write_text(
                json.dumps(report, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            if report["missing_paths"]:
                errors.append("revised ZIP is missing original archive files")
        except (OSError, ValueError, zipfile.BadZipFile) as error:
            errors.append(f"archive preservation verification failed: {error}")
print(json.dumps({"valid": not errors, "errors": errors, "files": [p.name for p in files]}, ensure_ascii=False, indent=2))
raise SystemExit(1 if errors else 0)
